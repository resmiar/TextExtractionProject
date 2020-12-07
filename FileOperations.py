import openpyxl
import xlsxwriter
import pickle


def write_file(tags_list):
    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook('Tags_list.xlsx')
    worksheet = workbook.add_worksheet()

    # Start from the first cell. Rows and columns are zero indexed.
    row = 0
    col = 0
    print("Writing to file")

    # Iterate over the data and write it out row by row.
    for attribute in tags_list.keys():
        worksheet.write(row, col, attribute)
        worksheet.write(row+1, col, tags_list[attribute])
        col += 1

    workbook.close()


def read_templates():
    with open('templates.txt', 'rb') as handle:
        templates_dict = dict()
        try:
            templates_dict = pickle.loads(handle.read())
        except EOFError as e:
            print("File is empty")
        print('Read: ', templates_dict)
        return templates_dict


def write_templates(templates_dict):
    with open('templates.txt', 'wb') as handle:
        pickle.dump(templates_dict, handle)
        print('write: ', templates_dict)


def write_bulk(tags, valueTag):
    # Creating work book
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Iterating tag list for header rows
    for index, key in enumerate(tags):
        sheet.cell(row=1, column=index+2).value = key
        print("headerValue:"+key)

    file_names = list(valueTag.keys())
    # Iterating dictionary for tag values
    for i, x in enumerate(valueTag.values()):
        file_name = file_names[i]
        print("Value of x:"+str(x))
        print('Filename is {}'.format(file_name))
        sheet.cell(row=i+2, column=1).value = file_name
        for idx,value in enumerate(x):
            sheet.cell(row=i+2, column=idx+2).value = value
            print("Value:"+value)

    # Saving workbook
    wb.save('tag.xlsx')